import streamlit as st
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import tempfile

st.set_page_config(page_title="Comparateur Excel", layout="wide")

st.title("🔎 Comparateur de fichiers Excel - Contrôle Interne")

st.write("Uploader deux fichiers Excel à comparer.")

file1 = st.file_uploader("📂 Fichier 1", type=["xlsx"])
file2 = st.file_uploader("📂 Fichier 2", type=["xlsx"])

if file1 and file2:

    try:
        df1 = pd.read_excel(file1)
        df2 = pd.read_excel(file2)

        # ===============================
        # 🔍 Vérifications structurelles
        # ===============================

        if df1.shape[1] != df2.shape[1]:
            st.error("❌ Les fichiers n'ont pas le même nombre de colonnes.")
            st.stop()

        if list(df1.columns) != list(df2.columns):
            st.error("❌ Les noms des colonnes sont différents.")
            st.write("Colonnes Fichier 1 :", list(df1.columns))
            st.write("Colonnes Fichier 2 :", list(df2.columns))
            st.stop()

        if df1.shape[0] != df2.shape[0]:
            st.error("❌ Les fichiers n'ont pas le même nombre de lignes.")
            st.stop()

        st.success("✅ Structure des fichiers valide. Comparaison en cours...")

        # ===============================
        # 🔎 Comparaison
        # ===============================

        df_result = df1.copy()
        differences_count = 0
        diff_by_column = {}

        for row in range(df1.shape[0]):
            for col in range(df1.shape[1]):

                val1 = df1.iloc[row, col]
                val2 = df2.iloc[row, col]

                if pd.isna(val1) and pd.isna(val2):
                    continue

                if val1 != val2:
                    differences_count += 1
                    col_name = df1.columns[col]
                    diff_by_column[col_name] = diff_by_column.get(col_name, 0) + 1

                    df_result.iloc[row, col] = f"FILE1 : {val1} / FILE2 : {val2}"

        # ===============================
        # 📊 Résumé
        # ===============================

        st.subheader("📊 Résumé des écarts")

        st.write(f"Nombre total de cellules différentes : **{differences_count}**")

        if differences_count > 0:
            st.write("Écarts par colonne :")
            st.json(diff_by_column)

        # ===============================
        # 💾 Génération fichier résultat
        # ===============================

        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
            temp_path = tmp.name
            df_result.to_excel(temp_path, index=False)

        wb = load_workbook(temp_path)
        ws = wb.active

        yellow_fill = PatternFill(start_color="FFFF00",
                                  end_color="FFFF00",
                                  fill_type="solid")

        for row in ws.iter_rows(min_row=2):
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("FILE1"):
                    cell.fill = yellow_fill

        wb.save(temp_path)

        st.success("🎉 Comparaison terminée avec succès !")


        file1_processed = file1.name.replace(".xlsx","_")
        file2_processed = file2.name.replace(".xlsx","_")

        with open(temp_path, "rb") as f:
            st.download_button(
                "📥 Télécharger le fichier comparé",
                f,
                file_name = r'Comparaison_'+file1_processed+file2_processed+'.xlsx'
            )

    except Exception as e:
        st.error(f"❌ Erreur lors du traitement : {str(e)}")